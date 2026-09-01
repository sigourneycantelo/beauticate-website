#!/usr/bin/env python3
"""
Builds the TGA review spreadsheet from docs/audit/lines-to-fix.json.

Every row in lines-to-fix.json was read by hand and classified below. Rows not
in DECISIONS were judged false positives and go on the second sheet, so the
filtering is checkable rather than something you have to take on trust.

  ACTION     a real issue under the Advertising Code. Suggested fix supplied.
  JUDGEMENT  arguably caught, but it needs an editor's call, not a rule.
  OK         flagged, reviewed, judged fine. Moved to the screened-out sheet
             with the reason, so the call is visible rather than silent.

Rationale and sources: docs/ask-sig-compliance.md
Output: docs/audit/tga-review.xlsx
"""

import json, pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = pathlib.Path(__file__).resolve().parent.parent
SRC = ROOT / "docs" / "audit" / "lines-to-fix.json"
OUT = ROOT / "docs" / "audit" / "tga-review.xlsx"

A, J, OK = "ACTION", "JUDGEMENT", "OK"

# OK rows were flagged by the detector, reviewed, and judged fine. They go on
# the screened-out sheet with their reason, rather than silently disappearing.
#
# The line that matters for interview routines: the Code catches *advertising a
# therapeutic good*. A generic substance with no brand and no claim ("she takes
# magnesium and B vitamins") names no product and promotes no supply, so there
# is nothing being advertised. It becomes a problem when a BRAND appears (a
# specific good) or a CLAIM appears (a therapeutic use), or when we sell it.

# index -> (verdict, suggested fix, note)
DECISIONS = {
    # ---- Sig's own personal use of therapeutic goods (sunscreen, devices) ----
    2:  (A, "The SPF we rate at the moment is mesoestetic.",
         "Sunscreens are therapeutic goods in Australia, so 'my hero SPF' is a testimonial."),
    3:  (A, "It's a beautifully elegant, high-protection sunscreen that sits perfectly under makeup and never leaves that heavy white cast, which is why it earns its place.",
         "Same sentence, same issue: 'the one I reach for most' is a personal use claim."),
    131:(A, "These are the sunscreens that earn their place, and the ones Lulu will tolerate.",
         "'The sunscreens I trust... I'll happily use on myself' is a testimonial about a therapeutic good."),
    132:(A, "The ones worth knowing about:", "Same article, same issue."),
    138:(J, "Optional: 'The Nook Infrared Sauna is one we rate, and the code BEAUTICATE gets you a discount.'",
         "DOWNGRADED. A sauna is very unlikely to be a therapeutic good, so the testimonial ban probably does not reach it. Caught only because 'infrared' is in the restricted list. Leave it unless the copy starts making health claims."),
    139:(A, "It suits relaxation, recovery and downtime.", "Personal benefit claim about a wellness device."),
    140:(A, "It's simple, accessible and easy to integrate into daily life.", "Drops 'I use it regularly'."),
    157:(A, "We've had Qure's LED tech in for testing for months, and the ritual is a lovely one.",
         "'The results' is the part that has to go: it is an efficacy claim, and it sits beside an affiliate link. The 'I've been testing' half is weaker and only bites if the device is a therapeutic good."),
    158:(J, "Optional: 'There's also a neck LED device worth knowing about.'",
         "DOWNGRADED. 'The bonus detail I love' is only a problem if the device is a therapeutic good, which I have not established for Qure. Same over-inclusion as the sauna row."),
    180:(A, "For me, I want to work on fine lines and pigmentation, so I use the Qure app to select the red light Anti-Aging Treatment.",
         "ADDED after review, I had screened this out. Naming your own skin condition and then the device setting you use for it is a personal use claim tied to a condition, which is the shape that matters most."),
    165:(J, "Optional: 'It's where the panel lives, beside the body oils, incense and fragrances.'",
         "DOWNGRADED. Depends on whether that specific panel is ARTG-listed, which I have not verified. The CLAIM rows in this same article are the ones that matter regardless."),
    178:(A, "Then the in-clinic tweakments, the at-home devices, and what we stock on the ingestibles side.",
         "Article intro. I missed this one in the first pass over this piece."),
    189:(A, "For every day the Moo Goo Invisible Zinc for Kids is a good option, and for the beach the La Roche Posay Anthelios XL Wet Skin Sunscreen is a clever innovation.",
         "Sunscreen is a therapeutic good. Advertising to children is separately restricted, so worth care here."),
    136:(A, "The facial finishes with the Healite II LED.", "Personal-use framing of a clinic device."),
    166:(J, "## What It's For", "Only if the Ultrahuman ring counts as a device. A fitness tracker probably does not."),

    # ---- LED / red light therapeutic claims. Biggest and most serious cluster ----
    186:(A, "REMOVE the section entirely.",
         "'How an LED Mask Helped Heal Burns'. Burns are a serious condition and this is the strongest claim found anywhere on the site."),
    155:(A, "DELETE the sentence.",
         "'Proven to speed up recovery and reduce inflammation in skin trauma and post-op healing' is a therapeutic claim about surgical recovery."),
    144:(A, "REMOVE the 'Wound Healing & Immunity' section.", "Wound healing is a therapeutic indication."),
    184:(A, "DELETE the sentence.", "'LED masks actually help reverse the damage of the sun' is a treatment claim."),
    154:(A, "DELETE the sentence.", "'Studies show it can improve firmness, reduce wrinkles, diminish pigmentation, and soothe inflammation.'"),
    32: (A, "The San Lueur Advanced LED Light Therapy Facial Mask offers three wavelengths: 415nm blue, 633nm red and 830nm near-infrared.",
         "Shop product. 'Clinically approved' and 'target breakouts' are both prohibited."),
    17: (A, "Red and near-infrared LED light therapy is used for skin tone and texture. Follow the manufacturer's directions for use.",
         "'Clinically shown to stimulate collagen production, reduce inflammation'."),
    147:(A, "Red light is used for skin tone and texture.", "Body copy version of the same claim."),
    19: (A, "The device uses red and near-infrared wavelengths.",
         "'Clinically proven wavelengths... the same ones dermatologists use' invokes practitioners as well."),
    153:(A, "This hands-free, at-home device uses two wavelengths, red light (633nm) and near-infrared (830nm).",
         "The strongest of the Qure rows. 'Clinically proven' plus 'to treat fine lines, pigmentation, redness' is an explicit treatment claim, made by us, beside an affiliate link. It needs substantiating under Australian Consumer Law whether or not the device is TGA-regulated."),
    25: (A, "LED masks like San Lueur have red and blue light settings, and microcurrent tools like NuFACE are often used before an event. Follow each manufacturer's directions for use.",
         "FAQ answer, so it feeds search results. 'Target breakout-causing bacteria' is an acne treatment claim."),
    26: (A, "It offers Red, Deep Red, Infrared, Amber and Blue light settings, with separate anti-ageing and blemish programmes.",
         "'Boost collagen and elastin production'."),
    31: (A, "These devices are designed for regular use over time. Follow the manufacturer's directions.",
         "'Boosts collagen production, reduces inflammation and improves skin texture'."),
    185:(A, "DELETE the sentence.", "Body copy version of the same claim."),
    30: (A, "consider an LED light therapy device such as the Dr Gross Spectralite Faceware Pro",
         "Drops 'to boost collagen production'."),
    63: (A, "REMOVE the Neutrogena mask reference entirely.",
         "'Clinically-proven... to stimulate collagen'. That specific mask was also subject to a TGA safety alert over retinal risk, so I would not reword it."),
    156:(A, "### Dual Lens LEDs", "'Medical-grade' is a claim we cannot substantiate."),
    122:(A, "Qure Q-Renew helmet with red LED light for the scalp",
         "Image alt text. 'Medical-grade LLLT lasers for scalp treatment'."),

    # ---- Clinic and directory listings (AHPRA territory) ----
    10: (A, "The clinic also offers yellow LED light therapy.", "'Promotes wound healing' about a clinic service."),
    12: (A, "The clinic offers LED light treatments as part of its menu.", "'Improves healing time'."),
    127:(A, "The clinic offers LED light treatments as part of its menu.", "'Subsiding active acne' is a treatment claim about a regulated health service."),
    125:(A, "LED Light Therapy. Ask the clinic which wavelengths they use and what the treatment involves.",
         "'Calm inflammation, stimulate collagen or combat acne'."),
    126:(A, "This is a light therapy treatment (which did sting a little). Ask the clinic what it involves and whether it suits you.",
         "'Treating broken capillaries' is a treatment claim."),
    8:  (J, "Replace 'medical-grade skincare' with the clinic's own wording.", "'Medical-grade' recurs across directory listings. Worth one consistent decision."),
    11: (J, "Replace 'medical-grade skin treatments' with the clinic's own wording.", "Same."),
    14: (J, "Replace 'medical-grade LED' with 'professional LED'.", "Same."),
    23: (J, "Replace 'medical-grade facial' with the clinic's own wording.", "Same."),
    128:(J, "Replace 'medical-grade LED' with 'professional LED'.", "Same."),
    142:(J, "Describe the device without the grade claim.", "Clinic equipment description."),

    # ---- Third-party supplement testimonials in interviews ----
    55: (A, "She takes collagen every morning.",
         "KEPT, but only the claim goes. Taking collagen is fine to report; 'which she credits with strengthening her previously brittle nails' is an efficacy claim about a supplement."),
    105:(A, "DELETE the sentence.",
         "KEPT as ACTION. This one has both problems: named brands (Viviscal, Lipocils) and therapeutic claims ('for hair growth', 'help regrow lashes')."),
    115:(A, "She takes zinc, vitamin C, magnesium and vitamin B shots.",
         "KEPT, but only the indications go. 'For the brain' and 'for immunity' state therapeutic uses; the substances themselves are fine."),
    88: (J, "Optional: drop the B12 shots reference and leave the rest.",
         "DOWNGRADED. Generic substances, no brand, no claim, so mostly fine. The only wrinkle is B12 injections, which are prescription-only in Australia, and prescription medicines cannot be advertised to the public at all."),
    93: (A, "She avoids red meat where possible, drinks apple cider vinegar and takes magnesium and zinc before bed.",
         "KEPT, but only the indication goes. 'To help with healing, relaxation' is the claim; the substances can stay."),
    98: (OK, "", "No change needed. Generic substances, no brand, no claim, not something we sell. Nothing is being advertised."),
    117:(OK, "", "No change needed. Generic substances, no brand, no claim."),
    76: (A, "She does Pilates twice a week and has a full health check every two years.",
         "KEPT as ACTION because The Super Elixir is a named brand, which makes it a specific good rather than a generic substance."),
    116:(OK, "", "No change needed. 'Herbal supplements' names no product and makes no claim."),
    162:(OK, "", "No change needed. Generic substances, no brand, no claim, and they are the interviewee's own words about her own routine."),
    62: (A, "She also has an infrared sauna at home.", "Drops 'for detoxification, relaxation and collagen production'."),
    73: (J, "Consider deleting.", "'Noticed a significant improvement in her skin since starting this routine' is a results claim, though about a routine rather than a named good."),

    # ---- Supplement and ingredient efficacy claims ----
    79: (A, "Joseph Hkeik says sunscreen is the number one essential.",
         "Claiming a product prevents skin cancer is a RESTRICTED REPRESENTATION under the Code and needs TGA approval. One of the most serious findings here."),
    84: (A, "Terri Vinson states that UV protection is the most important skincare product.", "Same restricted representation."),
    86: (A, "Vitamin B, also known as niacinamide, is used in skincare for moisture, barrier support and the appearance of pigmentation.",
         "'Can even treat acne' is a therapeutic claim."),
    24: (A, "Hydrolysed collagen and hyaluronic acid are common ingredients in ingestible beauty products. Always read the label and follow the directions for use.",
         "'Studies show... can improve skin elasticity and hydration'."),
    34: (A, "Comvita Olive Leaf Extract is available in liquid, lozenge or capsule form.",
         "Drops 'antioxidant, anti-inflammatory, antiviral' claims about a supplement."),
    187:(A, "Trim the quote to remove the therapeutic properties list, or drop the quote.",
         "Quoting an expert does not transfer responsibility. We publish it, we own it."),
    44: (A, "Jessica highlights selenium, magnesium, vitamin C and vitamin E.", "Drops 'help prevent early signs of ageing'."),
    71: (A, "Lucy discovered sea buckthorn capsules, which contain omega-7.", "Drops 'highly anti-inflammatory'."),
    188:(A, "Manuka honey is another popular natural remedy.", "Drops sore throats, digestive issues and wound healing."),
    172:(A, "Plant-based options like phytoestrogens (found in soy, flaxseed and legumes) are often discussed as alternatives. Speak with your GP about what suits you.",
         "'Can help balance hormones' plus an HRT comparison."),
    70: (A, "When the bacteria on your skin are in balance it can help keep skin looking calm and comfortable.",
         "Names acne and psoriasis alongside a product."),
    102:(J, "Kombucha is a fermented tea with a devoted following.", "'Prevent illness, detox the system'. Food rather than a therapeutic good, so your call."),
    50: (J, "Consider 'use a targeted gel'.", "'Anti-inflammatory gel to reduce swelling'."),
    37: (J, "Consider retitling.", "'How to Prevent Hair Loss'. Hair loss is a condition."),
    91: (J, "Consider rewording the FAQ question.", "'What's the best skincare routine for treating acne?' feeds FAQ schema."),
    92: (J, "Consider removing this FAQ.", "'Can the contraceptive pill cure adult acne?' concerns a prescription medicine, which cannot be advertised to the public at all."),
    148:(J, "Attribute plainly or drop the health claim.", "'Magnesium pools packed with healing properties', quoted from a supplier."),
    149:(J, "Same as above.", "Duplicate of the pool claim."),
    150:(J, "Consider dropping the quote.", "'You want a pool that heals, not just cools.'"),
    41: (J, "Consider trimming to the product recommendation.", "Practitioner named alongside 'prevent pigmentation'."),
    94: (J, "Consider trimming the claim.", "Third party saying light therapy 'heals'."),
}

rows = json.loads(SRC.read_text())

HEAD = ["#", "Priority", "Verdict", "Article", "Where", "Issue", "Date",
        "Offending line", "Suggested fix", "Approve Y/N", "Notes", "File"]

WIDTHS = [5, 9, 11, 34, 12, 13, 10, 62, 62, 12, 52, 60]

thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
head_fill = PatternFill("solid", fgColor="2F3E46")
action_fill = PatternFill("solid", fgColor="FDE7E9")
judge_fill = PatternFill("solid", fgColor="FFF6E5")


def write_sheet(ws, data, with_fix=True):
    ws.append(HEAD)
    for i, c in enumerate(ws[1], start=1):
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = head_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[c.column_letter].width = WIDTHS[i - 1]
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26

    for idx, r, verdict, fix, note in data:
        ws.append([
            idx,
            f"P{r['priority']}",
            verdict,
            r["title"][:90],
            r["zone"],
            r["kind"],
            (r["date"] or "")[:10],
            r["sentence"][:600],
            fix,
            "",
            note,
            r["file"],
        ])
        row = ws.max_row
        for c in ws[row]:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border
        if verdict == "ACTION":
            ws.cell(row=row, column=3).fill = action_fill
        elif verdict == "JUDGEMENT":
            ws.cell(row=row, column=3).fill = judge_fill
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.row_dimensions[row].height = 58

    if with_fix and ws.max_row > 1:
        dv = DataValidation(type="list", formula1='"Y,N,LATER"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"J2:J{ws.max_row}")


action, screened = [], []
for i, r in enumerate(rows):
    if i in DECISIONS:
        v, fix, note = DECISIONS[i]
        if v == OK:
            screened.append((i, r, "reviewed - OK", "", note))
        else:
            action.append((i, r, v, fix, note))
    else:
        screened.append((i, r, "false positive", "", "Screened out on review. Listed so you can check the filtering."))

# ACTION first, then JUDGEMENT; most recent first inside each.
action.sort(key=lambda x: (x[2] != "ACTION", -(int((x[1]["date"] or "0")[:4] or 0))))

wb = Workbook()
ws1 = wb.active
ws1.title = "Action needed"
write_sheet(ws1, action)

ws2 = wb.create_sheet("Screened out")
write_sheet(ws2, screened, with_fix=False)

wb.save(OUT)

n_act = sum(1 for a in action if a[2] == "ACTION")
n_jud = len(action) - n_act
print(f"Action needed : {len(action)} rows  ({n_act} ACTION, {n_jud} JUDGEMENT)")
print(f"Screened out  : {len(screened)} rows")
print(f"Articles to edit: {len(set(a[1]['file'] for a in action))}")
print(f"→ {OUT.relative_to(ROOT)}")
